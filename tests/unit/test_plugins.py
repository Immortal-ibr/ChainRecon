"""Comprehensive tests for the plugin system."""

import csv
import json
import tempfile
import unittest
from pathlib import Path

from plugins import CsvExportPlugin, HtmlReportPlugin, JsonReportPlugin, XlsxReportPlugin, get_plugin
from plugins.base import ReportPlugin
from openpyxl import load_workbook


SAMPLE_DATA = {
    "traffic": {
        "findings": {
            "dns_queries": [
                {"query": "api.vendor.example", "src_ip": "192.168.1.10"},
                {"query": "telemetry.cloud.io", "src_ip": "192.168.1.10"},
            ],
            "external_ips": ["93.184.216.34", "1.2.3.4"],
            "http_requests": [{"method": "GET", "host": "device.local", "uri": "/status"}],
        },
        "summary": {"dns_query_count": 2},
    },
    "ssl": {
        "findings": {
            "certificates": [
                {"port": 443, "subject": "CN=device", "self_signed": True}
            ],
            "security_findings": [{"severity": "medium", "title": "Self-signed cert"}],
        },
    },
    "scan": {
        "findings": {
            "hosts": [
                {"ip": "192.168.1.50", "services": [{"port": 80, "service": "http"}, {"port": 1883, "service": "mqtt"}]}
            ],
            "iot_services": [{"port": 1883, "service": "mqtt", "protocol": "MQTT"}],
        },
    },
}


# ===========================================================================
# Base plugin contract tests
# ===========================================================================


class PluginContractTests(unittest.TestCase):
    """Verify all plugins satisfy the ReportPlugin ABC."""

    def test_json_plugin_is_report_plugin(self):
        self.assertIsInstance(JsonReportPlugin(), ReportPlugin)

    def test_html_plugin_is_report_plugin(self):
        self.assertIsInstance(HtmlReportPlugin(), ReportPlugin)

    def test_csv_plugin_is_report_plugin(self):
        self.assertIsInstance(CsvExportPlugin(), ReportPlugin)

    def test_xlsx_plugin_is_report_plugin(self):
        self.assertIsInstance(XlsxReportPlugin(), ReportPlugin)

    def test_plugins_have_name_attribute(self):
        self.assertEqual(JsonReportPlugin.name, "json")
        self.assertEqual(HtmlReportPlugin.name, "html")
        self.assertEqual(CsvExportPlugin.name, "csv")
        self.assertEqual(XlsxReportPlugin.name, "xlsx")

    def test_plugins_have_description(self):
        for cls in (JsonReportPlugin, HtmlReportPlugin, CsvExportPlugin, XlsxReportPlugin):
            self.assertTrue(len(cls.description) > 0, f"{cls.__name__} has no description")

    def test_plugins_have_file_extension(self):
        self.assertEqual(JsonReportPlugin().file_extension(), ".json")
        self.assertEqual(HtmlReportPlugin().file_extension(), ".html")
        self.assertEqual(CsvExportPlugin().file_extension(), ".csv")
        self.assertEqual(XlsxReportPlugin().file_extension(), ".xlsx")


# ===========================================================================
# Registry tests
# ===========================================================================


class RegistryTests(unittest.TestCase):
    def test_lookup_json(self):
        self.assertIsInstance(get_plugin("json"), JsonReportPlugin)

    def test_lookup_html(self):
        self.assertIsInstance(get_plugin("html"), HtmlReportPlugin)

    def test_lookup_csv(self):
        self.assertIsInstance(get_plugin("csv"), CsvExportPlugin)

    def test_lookup_xlsx(self):
        self.assertIsInstance(get_plugin("xlsx"), XlsxReportPlugin)

    def test_lookup_case_insensitive(self):
        self.assertIsInstance(get_plugin("JSON"), JsonReportPlugin)
        self.assertIsInstance(get_plugin("Html"), HtmlReportPlugin)

    def test_unknown_plugin_raises_value_error(self):
        with self.assertRaises(ValueError):
            get_plugin("xml")

    def test_empty_name_raises(self):
        with self.assertRaises((ValueError, KeyError)):
            get_plugin("")


# ===========================================================================
# JSON plugin tests
# ===========================================================================


class JsonPluginTests(unittest.TestCase):
    def test_writes_valid_json(self):
        with tempfile.NamedTemporaryFile("r+", suffix=".json", delete=False, encoding="utf-8") as f:
            path = JsonReportPlugin().generate(SAMPLE_DATA, f.name)
            f.seek(0)
            payload = json.load(f)
        self.assertEqual(path, f.name)
        self.assertIn("traffic", payload)
        self.assertIn("ssl", payload)
        self.assertIn("scan", payload)

    def test_keys_are_sorted(self):
        with tempfile.NamedTemporaryFile("r+", suffix=".json", delete=False, encoding="utf-8") as f:
            JsonReportPlugin().generate(SAMPLE_DATA, f.name)
            f.seek(0)
            content = f.read()
        # "scan" should appear before "ssl" before "traffic" in sorted keys
        scan_pos = content.index('"scan"')
        ssl_pos = content.index('"ssl"')
        traffic_pos = content.index('"traffic"')
        self.assertLess(scan_pos, ssl_pos)
        self.assertLess(ssl_pos, traffic_pos)

    def test_output_is_indented(self):
        with tempfile.NamedTemporaryFile("r+", suffix=".json", delete=False, encoding="utf-8") as f:
            JsonReportPlugin().generate(SAMPLE_DATA, f.name)
            f.seek(0)
            content = f.read()
        self.assertIn("\n", content)
        self.assertIn("  ", content)

    def test_returns_output_path(self):
        with tempfile.NamedTemporaryFile("w", suffix=".json", delete=False, encoding="utf-8") as f:
            result = JsonReportPlugin().generate({"a": None}, f.name)
        self.assertEqual(result, f.name)

    def test_handles_none_sections(self):
        data = {"traffic": None, "ssl": None, "scan": None}
        with tempfile.NamedTemporaryFile("r+", suffix=".json", delete=False, encoding="utf-8") as f:
            JsonReportPlugin().generate(data, f.name)
            f.seek(0)
            payload = json.load(f)
        self.assertIsNone(payload["traffic"])

    def test_handles_empty_data(self):
        with tempfile.NamedTemporaryFile("r+", suffix=".json", delete=False, encoding="utf-8") as f:
            JsonReportPlugin().generate({}, f.name)
            f.seek(0)
            payload = json.load(f)
        self.assertEqual(payload, {})


# ===========================================================================
# HTML plugin tests
# ===========================================================================


class HtmlPluginTests(unittest.TestCase):
    def _generate(self, data=None):
        data = data or SAMPLE_DATA
        with tempfile.NamedTemporaryFile("r+", suffix=".html", delete=False, encoding="utf-8") as f:
            HtmlReportPlugin().generate(data, f.name)
            f.seek(0)
            return f.read(), f.name

    def test_contains_title(self):
        content, _ = self._generate()
        self.assertIn("ChainRecon Analysis Report", content)

    def test_contains_html_structure(self):
        content, _ = self._generate()
        self.assertIn("<!doctype html", content.lower())
        self.assertIn("</html>", content.lower())

    def test_contains_section_data(self):
        content, _ = self._generate()
        self.assertIn("api.vendor.example", content)

    def test_uses_collapsible_sections(self):
        content, _ = self._generate()
        self.assertIn("<details", content)
        self.assertIn("Toggle section", content)

    def test_renders_file_links_for_source_files(self):
        data = {
            "traffic": {
                "metadata": {"source_file": "/tmp/example.json"},
                "findings": {},
                "summary": {},
            }
        }
        content, _ = self._generate(data)
        self.assertIn("file:///", content)

    def test_escapes_special_characters(self):
        """Ensure user-controlled data is escaped to prevent XSS."""
        data = {"xss_test": {"findings": {"payload": "<script>alert('xss')</script>"}}}
        content, _ = self._generate(data)
        self.assertNotIn("<script>alert", content)

    def test_handles_none_sections(self):
        data = {"traffic": None, "ssl": None}
        content, _ = self._generate(data)
        self.assertIn("ChainRecon", content)

    def test_returns_output_path(self):
        with tempfile.NamedTemporaryFile("w", suffix=".html", delete=False, encoding="utf-8") as f:
            result = HtmlReportPlugin().generate(SAMPLE_DATA, f.name)
        self.assertEqual(result, f.name)


# ===========================================================================
# CSV plugin tests
# ===========================================================================


class CsvPluginTests(unittest.TestCase):
    def _generate(self, data=None):
        data = data or SAMPLE_DATA
        with tempfile.NamedTemporaryFile("r+", suffix=".csv", delete=False, encoding="utf-8", newline="") as f:
            CsvExportPlugin().generate(data, f.name)
            f.seek(0)
            return list(csv.DictReader(f)), f.name

    def test_flattens_traffic_findings(self):
        rows, _ = self._generate()
        traffic_rows = [r for r in rows if r.get("section") == "traffic"]
        self.assertGreater(len(traffic_rows), 0)

    def test_flattens_scan_findings(self):
        rows, _ = self._generate()
        scan_rows = [r for r in rows if r.get("section") == "scan"]
        self.assertGreater(len(scan_rows), 0)

    def test_all_rows_have_section_and_key(self):
        rows, _ = self._generate()
        for row in rows:
            self.assertIn("page", row)
            self.assertIn("section", row)
            self.assertIn("key", row)

    def test_handles_all_none_sections(self):
        data = {"traffic": None, "ssl": None, "scan": None}
        rows, _ = self._generate(data)
        self.assertEqual(len(rows), 0)

    def test_handles_scalar_findings(self):
        data = {"test": {"findings": {"simple_value": "hello"}}}
        rows, _ = self._generate(data)
        matching = [r for r in rows if r.get("key") == "simple_value"]
        self.assertEqual(len(matching), 1)
        self.assertEqual(matching[0]["value"], "hello")

    def test_handles_dict_findings(self):
        data = {"test": {"findings": {"info": {"a": "1", "b": "2"}}}}
        rows, _ = self._generate(data)
        matching = [r for r in rows if r.get("key") == "info"]
        self.assertEqual(len(matching), 1)
        self.assertEqual(matching[0]["a"], "1")

    def test_handles_list_of_scalars(self):
        data = {"test": {"findings": {"ips": ["1.2.3.4", "5.6.7.8"]}}}
        rows, _ = self._generate(data)
        matching = [r for r in rows if r.get("key") == "ips"]
        self.assertEqual(len(matching), 2)

    def test_has_header_row(self):
        with tempfile.NamedTemporaryFile("r+", suffix=".csv", delete=False, encoding="utf-8", newline="") as f:
            CsvExportPlugin().generate(SAMPLE_DATA, f.name)
            f.seek(0)
            reader = csv.reader(f)
            header = next(reader)
        self.assertIn("section", header)
        self.assertIn("key", header)

    def test_returns_output_path(self):
        with tempfile.NamedTemporaryFile("w", suffix=".csv", delete=False, encoding="utf-8", newline="") as f:
            result = CsvExportPlugin().generate(SAMPLE_DATA, f.name)
        self.assertEqual(result, f.name)


class XlsxPluginTests(unittest.TestCase):
    def test_writes_summary_and_section_sheets(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = XlsxReportPlugin().generate(SAMPLE_DATA, f.name)
        workbook = load_workbook(path)
        self.assertIn("summary", workbook.sheetnames)
        self.assertIn("traffic", workbook.sheetnames)
        self.assertIn("ssl", workbook.sheetnames)
        self.assertIn("scan", workbook.sheetnames)
        summary = workbook["summary"]
        rows = list(summary.iter_rows(values_only=True))
        traffic_row = next(row for row in rows if row and row[0] == "traffic")
        self.assertGreater(int(traffic_row[1]), 0)
        workbook.close()
        Path(path).unlink(missing_ok=True)

    def test_csv_and_html_render_frida_sessions(self):
        frida_data = {
            "frida": {
                "metadata": {"section": "frida"},
                "sessions": [
                    {
                        "target": "com.nooie.home",
                        "script": "hook_all_methods",
                        "status": "stopped_by_user",
                        "hook_event_count": 3,
                        "error_count": 1,
                        "summary": {"status": "stopped_by_user"},
                        "events_by_tag": {"HOOK": 3, "WARN": 1},
                        "hook_events": ["[HOOK] java.net.Socket.connect"],
                        "error_events": ["[WARN] retrying"],
                        "artifacts": [{"type": "frida_summary", "path": "/tmp/session.json"}],
                    }
                ],
                "findings": {"sessions": []},
                "summary": {"session_count": 1},
                "risk_indicators": [],
                "artifacts": [],
            }
        }
        with tempfile.NamedTemporaryFile("r+", suffix=".csv", delete=False, encoding="utf-8", newline="") as csv_file:
            CsvExportPlugin().generate(frida_data, csv_file.name)
            csv_file.seek(0)
            csv_text = csv_file.read()
        self.assertIn("session_hook_event", csv_text)
        with tempfile.NamedTemporaryFile("r+", suffix=".html", delete=False, encoding="utf-8") as html_file:
            HtmlReportPlugin().generate(frida_data, html_file.name)
            html_file.seek(0)
            html_text = html_file.read()
        self.assertIn("Frida Sessions", html_text)
        self.assertIn("java.net.Socket.connect", html_text)


if __name__ == "__main__":
    unittest.main()
